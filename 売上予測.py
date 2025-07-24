import pandas as pd
from openpyxl import load_workbook
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import LabelEncoder
import math

excel_path = r'c:\Users\taiyo\OneDrive\デスクトップ\売り上げサンプル.xlsx'
sheet_name = '売上データ'

# データ読み込み
df = pd.read_excel(excel_path, sheet_name=sheet_name)

# カテゴリ変数を数値に変換
le_youbi = LabelEncoder()
le_tenki = LabelEncoder()
df['曜日区分'] = le_youbi.fit_transform(df['曜日区分'])
df['天気'] = le_tenki.fit_transform(df['天気'])

# 前回予測・前回実績がなければ0で埋める
if '前回予測' not in df.columns:
    df['前回予測'] = 0
if '前回実績' not in df.columns:
    df['前回実績'] = 0

# 特徴量とターゲット
X = df[['曜日区分', '天気', '前回予測', '前回実績']]
y = df['売上']

# モデル学習
model = LinearRegression()
model.fit(X, y)

# ユーザー入力
youbi = input('曜日区分を入力してください（平日/休日/祝祭日）：')
tenki = input('天気を入力してください（晴れ/雨/曇り）：')

# 入力値を変換
input_youbi = le_youbi.transform([youbi])
input_tenki = le_tenki.transform([tenki])

# 前回の予測・実績を取得（なければ0）
if len(df) > 0:
    prev_yoso = df.iloc[-1]['前回予測']
    prev_jissai = df.iloc[-1]['前回実績']
else:
    prev_yoso = 0
    prev_jissai = 0

# 予測
yoso = model.predict([[input_youbi[0], input_tenki[0], prev_yoso, prev_jissai]])[0]
yoso_100 = int(round(yoso / 100) * 100)
print(f'{youbi}・{tenki}の予測売上は {yoso_100} 円です。')

# 実際の売上を入力
jissai_uriage = int(input('実際の売上を入力してください（数字のみ）：'))

# Excelファイルに追記
wb = load_workbook(excel_path)
ws = wb[sheet_name]
ws.append([youbi, tenki, yoso_100, jissai_uriage, prev_yoso, prev_jissai])

# メニューシートに仕入れ数を記載（合計がyoso_100になるように）
menu_ws = wb['メニュー']
menu_header = [cell.value for cell in menu_ws[1]]
tanka_col = menu_header.index('単価') + 1
shiire_col = len(menu_header) + 1  # 新しい列（仕入れ数）

if '仕入れ数' not in menu_header:
    menu_ws.cell(row=1, column=shiire_col, value='仕入れ数')

# メニューごとの単価リスト
menu_rows = []
tanka_list = []
for row in range(2, menu_ws.max_row + 1):
    tanka = menu_ws.cell(row=row, column=tanka_col).value
    if tanka is not None and tanka != 0:
        tanka_list.append(tanka)
        menu_rows.append(row)
    else:
        tanka_list.append(0)
        menu_rows.append(row)

menu_count = len([t for t in tanka_list if t > 0])
if menu_count == 0:
    shiire_list = [0 for _ in tanka_list]
else:
    budget_per_menu = yoso_100 // menu_count
    shiire_list = [budget_per_menu // tanka if tanka > 0 else 0 for tanka in tanka_list]
    total = sum([shiire * tanka for shiire, tanka in zip(shiire_list, tanka_list)])
    nokori = yoso_100 - total
    # 残りの金額で最後のメニューを追加購入
    if tanka_list[-1] > 0:
        add = nokori // tanka_list[-1]
        shiire_list[-1] += add
        total += add * tanka_list[-1]

# 書き込み
for row, shiire in zip(menu_rows, shiire_list):
    menu_ws.cell(row=row, column=shiire_col, value=shiire)

wb.save(excel_path)
print(f'予測値 {yoso_100} 円・実際の売上 {jissai_uriage} 円 を{excel_path}に追加し、仕入れ数も更新しました。') 